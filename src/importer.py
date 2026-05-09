from __future__ import annotations

import io
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from src.formulas import safe_evaluate
from src.models import (
    Campaign,
    Character,
    CharacterDerived,
    CharacterResource,
    CharacterStat,
    Curse,
    InventoryItem,
    Pet,
    PetSkill,
    Quest,
    Ruleset,
    Skill,
    Talent,
)

STAT_KEY_RE = re.compile(r"\(([A-Z]{2,4})\)")


def _num(v: Any, default: float = 0.0) -> float:
    if v is None or v == "":
        return default
    if isinstance(v, (int, float)):
        return float(v)
    try:
        s = str(v).strip().replace(",", ".")
        if s.endswith("%"):
            s = s[:-1]
        return float(s)
    except (ValueError, TypeError):
        return default


def _str(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _ensure_default_ruleset(session: Session) -> Ruleset:
    rs = session.scalars(select(Ruleset)).first()
    if rs:
        return rs
    camp = session.scalars(select(Campaign)).first()
    if camp is None:
        camp = Campaign(name="Aethermoor")
        session.add(camp)
        session.flush()
    rs = Ruleset(campaign_id=camp.id, name="Default")
    session.add(rs)
    session.flush()
    return rs


def import_from_xlsx_bytes(
    session: Session,
    data: bytes,
    *,
    owner_id: int | None = None,
) -> Character:
    return _import(session, io.BytesIO(data), owner_id=owner_id)


def import_from_xlsx_path(
    session: Session,
    path: str | Path,
    *,
    owner_id: int | None = None,
) -> Character:
    return _import(session, str(path), owner_id=owner_id)


def _import(session: Session, source, *, owner_id: int | None) -> Character:
    wb = load_workbook(source, data_only=True)
    main = None
    for name in wb.sheetnames:
        if name.lower() not in {
            "skills", "skill attive", "inventario", "talenti",
            "pets", "maledizioni", "quest",
        }:
            main = wb[name]
            break
    if main is None:
        raise ValueError("Foglio principale non trovato.")

    rs = _ensure_default_ruleset(session)
    camp_id = rs.campaign_id

    name = _str(main["B4"].value) or main.title
    class_name = _str(main["D4"].value)
    level = int(_num(main["B5"].value, 1) or 1)
    xp = int(_num(main["D5"].value, 0))

    char = Character(
        name=name,
        class_name=class_name,
        level=level,
        xp=xp,
        owner_id=owner_id,
        campaign_id=camp_id,
    )
    session.add(char)
    session.flush()

    _import_stats(session, char, main)
    _import_resources_and_derived(session, char, main, rs)
    _import_titles(session, char, main)

    if "Skills" in wb.sheetnames:
        _import_skills(session, char, wb["Skills"], active=False)
    if "Skill Attive" in wb.sheetnames:
        _import_skills(session, char, wb["Skill Attive"], active=True)
    if "Talenti" in wb.sheetnames:
        _import_talents(session, char, wb["Talenti"])
    if "Pets" in wb.sheetnames:
        _import_pets(session, char, wb["Pets"])
    if "Inventario" in wb.sheetnames:
        _import_inventory(session, char, wb["Inventario"])
    if "Maledizioni" in wb.sheetnames:
        _import_curses(session, char, wb["Maledizioni"])
    if "Quest" in wb.sheetnames:
        _import_quests(session, char, wb["Quest"])

    session.flush()
    _refresh_resource_currents(char)
    session.flush()
    return char


def _import_stats(session: Session, char: Character, sheet) -> None:
    for row in range(9, 25):
        label = _str(sheet.cell(row=row, column=1).value)
        if not label or label.startswith("Statistica "):
            continue
        m = STAT_KEY_RE.search(label)
        key = m.group(1) if m else label[:6].upper().replace(" ", "_")
        clean_label = STAT_KEY_RE.sub("", label).strip()

        v_init = _num(sheet.cell(row=row, column=3).value)
        v_creat = _num(sheet.cell(row=row, column=4).value)
        v_inv = _num(sheet.cell(row=row, column=5).value)
        v_lvl = _num(sheet.cell(row=row, column=6).value)
        v_other = _num(sheet.cell(row=row, column=7).value)

        session.add(CharacterStat(
            character_id=char.id,
            key=key,
            label=clean_label,
            sort_order=row,
            value_initial=v_init,
            value_creation=v_creat,
            value_invested=v_inv,
            value_levelup=v_lvl,
            value_other=v_other,
        ))


def _import_titles(session: Session, char: Character, sheet) -> None:
    stat_keys = ["STR", "DEX", "VIT", "INT", "FIN", "AUT"]
    for row in range(37, 51):
        title_name = _str(sheet.cell(row=row, column=6).value)
        if not title_name:
            continue
        rarity = _str(sheet.cell(row=row, column=7).value)
        bonuses: dict[str, float] = {}
        for i, k in enumerate(stat_keys):
            v = _num(sheet.cell(row=row, column=8 + i).value)
            if v:
                bonuses[k] = v
        from src.models import Title
        session.add(Title(
            character_id=char.id, name=title_name, rarity=rarity,
            stat_bonuses_json=bonuses,
        ))


_DERIVED_FORMULAS = {
    "Moltiplicatore Danno Fisico": ("PHYS_DMG_MUL", "1 + STR / 100", "x"),
    "Capacita Carico (kg)": ("CARRY_KG", "50 + STR * 2", "kg"),
    "Capacità Carico (kg)": ("CARRY_KG", "50 + STR * 2", "kg"),
    "Attacchi per Turno": ("ATK_PER_TURN", "1 + DEX * 0.005", ""),
    "Precisione (%)": ("ACCURACY", "70 + DEX / 10", "%"),
    "Evasione (%)": ("EVASION", "DEX / 20", "%"),
    "Velocita Movimento (%)": ("MOVE_SPEED", "100 + DEX / 5", "%"),
    "Velocità Movimento (%)": ("MOVE_SPEED", "100 + DEX / 5", "%"),
    "Punti Vita (HP)": ("HP_MAX", "100 + VIT * 15", ""),
    "Rigenerazione (HP/min)": ("REGEN_HP", "VIT / 10", ""),
    "Resistenza Veleni (%)": ("RES_POISON", "VIT / 5", "%"),
    "Resistenza Elementi (%)": ("RES_ELEM", "VIT / 10", "%"),
    "Punti Mana (MP)": ("MP_MAX", "50 + INT * 10", ""),
    "Moltiplicatore Danno Magico": ("MAGIC_DMG_MUL", "1 + INT / 60", "x"),
    "Bonus EXP (%)": ("EXP_BONUS", "INT / 20", "%"),
    "Bonus Crafting (%)": ("CRAFT_BONUS", "(INT + FIN) / 10", "%"),
}


def _import_resources_and_derived(session: Session, char: Character, sheet, rs: Ruleset) -> None:
    derived_done: set[str] = set()
    for row in range(30, 44):
        label = _str(sheet.cell(row=row, column=1).value)
        if not label:
            continue
        bonus = _num(sheet.cell(row=row, column=3).value)

        meta = _DERIVED_FORMULAS.get(label)
        if meta is None:
            for k, v in _DERIVED_FORMULAS.items():
                if k.lower() == label.lower():
                    meta = v
                    break
        if meta is None:
            continue
        key, formula, unit = meta
        derived_done.add(key)

        if key == "HP_MAX":
            session.add(CharacterResource(
                character_id=char.id, key="HP", label="Punti Vita",
                color_hex="#dc2626", max_formula=formula,
                regen_formula="VIT / 10", current_value=0.0, sort_order=1,
            ))
        elif key == "MP_MAX":
            session.add(CharacterResource(
                character_id=char.id, key="MP", label="Mana",
                color_hex="#2563eb", max_formula=formula,
                regen_formula="0", current_value=0.0, sort_order=2,
            ))
        else:
            session.add(CharacterDerived(
                character_id=char.id, key=key, label=label,
                formula=formula, unit=unit, item_bonus=bonus, sort_order=row,
            ))

    for r_t in rs.resources:
        if not any(rr.key == r_t.key for rr in session.new if isinstance(rr, CharacterResource) and rr.character_id == char.id):
            existing = session.query(CharacterResource).filter_by(character_id=char.id, key=r_t.key).first() if char.id else None
            if existing is None:
                session.add(CharacterResource(
                    character_id=char.id, key=r_t.key, label=r_t.label,
                    color_hex=r_t.color_hex, max_formula=r_t.max_formula,
                    regen_formula=r_t.regen_formula, current_value=0.0,
                    sort_order=r_t.sort_order,
                ))


def _refresh_resource_currents(char: Character) -> None:
    names = {}
    titles_by_stat: dict[str, float] = {}
    for t in char.titles:
        for k, v in (t.stat_bonuses_json or {}).items():
            try:
                titles_by_stat[k] = titles_by_stat.get(k, 0.0) + float(v)
            except (TypeError, ValueError):
                continue
    for st_o in char.stats:
        names[st_o.key] = (
            st_o.value_initial + st_o.value_creation + st_o.value_invested
            + st_o.value_levelup + st_o.value_other
            + titles_by_stat.get(st_o.key, 0.0)
        )
    names["LEVEL"] = float(char.level)
    for r in char.resources:
        max_val, _ = safe_evaluate(r.max_formula, names)
        r.current_value = max_val


def _import_skills(session: Session, char: Character, sheet, *, active: bool) -> None:
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        name = _str(row[0])
        if not name or name.lower() == "domatore":
            continue
        if active:
            level = _num(row[2] if len(row) > 2 else 0)
            pct = _num(row[3] if len(row) > 3 else 0)
            desc = _str(row[1] if len(row) > 1 else "")
            cat = ""
            is_max = False
        else:
            level_raw = row[1] if len(row) > 1 else 0
            is_max = isinstance(level_raw, str) and level_raw.upper() == "MAX"
            level = _num(level_raw) if not is_max else 10.0
            pct_raw = row[2] if len(row) > 2 else 0
            pct = 100.0 if isinstance(pct_raw, str) and pct_raw.upper() == "MAX" else _num(pct_raw)
            desc = _str(row[3] if len(row) > 3 else "")
            cat = _str(row[4] if len(row) > 4 else "")
        session.add(Skill(
            character_id=char.id, name=name, level=level,
            progress_pct=pct, category=cat, description=desc,
            is_active=active, is_max=is_max,
        ))


def _import_talents(session: Session, char: Character, sheet) -> None:
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        name = _str(row[0])
        if not name or name == "--":
            continue
        rarity = _str(row[1] if len(row) > 1 else "")
        origin = _str(row[2] if len(row) > 2 else "")
        effect = _str(row[3] if len(row) > 3 else "")
        session.add(Talent(
            character_id=char.id, name=name, rarity=rarity,
            origin=origin, effect=effect, stat_bonuses_json={},
        ))


def _import_pets(session: Session, char: Character, sheet) -> None:
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        pet = Pet(
            character_id=char.id,
            name=_str(row[0]),
            level=int(_num(row[1] if len(row) > 1 else 1)),
            species=_str(row[2] if len(row) > 2 else ""),
            passive_skill=_str(row[3] if len(row) > 3 else ""),
        )
        session.add(pet)
        session.flush()
        for offset in range(4, min(len(row), 16), 3):
            sk_name = _str(row[offset])
            if not sk_name:
                continue
            sk_lvl = _num(row[offset + 1] if offset + 1 < len(row) else 0)
            sk_pct = _num(row[offset + 2] if offset + 2 < len(row) else 0)
            session.add(PetSkill(
                pet_id=pet.id, name=sk_name, level=sk_lvl, progress_pct=sk_pct,
            ))


def _import_inventory(session: Session, char: Character, sheet) -> None:
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        session.add(InventoryItem(
            character_id=char.id,
            name=_str(row[0]),
            stats_text=_str(row[1] if len(row) > 1 else ""),
            rarity=_str(row[2] if len(row) > 2 else ""),
            item_type=_str(row[3] if len(row) > 3 else ""),
            quantity=_num(row[4] if len(row) > 4 else 1, default=1.0),
            weight_kg=_num(row[5] if len(row) > 5 else 0),
        ))


def _import_curses(session: Session, char: Character, sheet) -> None:
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        session.add(Curse(
            character_id=char.id,
            name=_str(row[0]),
            description=_str(row[1] if len(row) > 1 else ""),
            bonus=_str(row[2] if len(row) > 2 else ""),
            malus=_str(row[3] if len(row) > 3 else ""),
        ))


def _import_quests(session: Session, char: Character, sheet) -> None:
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        session.add(Quest(
            character_id=char.id,
            name=_str(row[0]),
            progress=_str(row[1] if len(row) > 1 else ""),
            reward=_str(row[2] if len(row) > 2 else ""),
        ))
